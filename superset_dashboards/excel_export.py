"""Render a FetchResult into a styled .xlsx workbook (in-memory bytes)."""
from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from superset_client import FetchResult


# (header label, DashboardRow attribute, column width)
_COLUMNS = [
    ("Dashboard", "title", 42),
    ("Team", "team", 22),
    ("Owner(s)", "owners", 26),
    ("Frequency", "frequency", 22),
    ("Disposition", "disposition", 13),
    ("Status", "status", 11),
    ("Tags", "tags", 24),
    ("Roles", "roles", 20),
    ("Last modified", "last_modified", 16),
    ("Modified by", "last_modified_by", 20),
    ("Created", "created_on", 16),
    ("URL", "url", 50),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="111827")
_META_FONT = Font(italic=True, size=10, color="6B7280")
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_STRIPE = PatternFill("solid", fgColor="F3F4F6")


def build_workbook(result: FetchResult) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboards"

    ncols = len(_COLUMNS)
    last_col = get_column_letter(ncols)

    # Title band
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "Superset Dashboard Register"
    ws["A1"].font = _TITLE_FONT

    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        f"Source: {result.base_url}    "
        f"Dashboards: {result.count}    "
        f"Generated: {generated}"
    )
    ws["A2"].font = _META_FONT

    header_row = 4
    for c, (label, _attr, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=c, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[header_row].height = 20

    for i, row in enumerate(result.dashboards):
        r = header_row + 1 + i
        values = row.as_dict()
        for c, (_label, attr, _width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=values.get(attr, ""))
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(attr == "title"))
            if i % 2 == 1:
                cell.fill = _STRIPE

    # Freeze header + enable autofilter over the table
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(result.dashboards)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
